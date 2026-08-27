#!/usr/bin/env python3
"""Build consumption.xlsx from batches.jsonl — what each batch actually used.

  python3 consumption.py [batches.jsonl] [consumption.xlsx]

`batches.jsonl` is the source of truth: append-only, one line per batch, which
is the shape that survives a Pi losing power mid-write. This workbook is a view
of it, rebuilt from scratch each time. Delete it and it comes back correct; it
can never drift from the log, because nothing ever writes to it directly.

Grain is one row per ingredient per batch, which is what an inventory system
will want: sum a column to get how much of a material went out over any period.
"""

import argparse
import json
import os
import sys
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3552")
HEAD_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=14, bold=True, color="1F3552")
RED = Font(name=FONT, size=10, color="9C0006")
MUTED = Font(name=FONT, size=9, italic=True, color="666666")


def load(path):
    if not os.path.exists(path):
        return []
    out = []
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                out.append(json.loads(line))
            except json.JSONDecodeError:
                pass          # a torn final line from a power cut, not fatal
    return out


def header(ws, row, labels, widths=None):
    for i, text in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=text)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def build_consumption(wb, batches):
    ws = wb.create_sheet("Consumption")
    ws["A1"] = "Material used, per ingredient per batch"
    ws["A1"].font = TITLE
    ws["A2"] = ("One row per ingredient per batch. Rebuilt from batches.jsonl — "
                "edits here are lost on the next rebuild. 'assumed' means the "
                "figure is the recipe target, because the bench scale is not "
                "wired to the Pi and its reading never reached it.")
    ws["A2"].font = MUTED
    header(ws, 4,
           ["Date", "Batch", "Product", "Base", "Base weight (g)", "Ingredient",
            "Target (g)", "Actual (g)", "Deviation (g)", "Weighed on",
            "Measured?", "Verified", "Water ratio"],
           [12, 10, 16, 10, 14, 24, 11, 11, 13, 13, 12, 11, 11])

    row = 5
    for b in batches:
        day = b.get("production_day") or (b.get("logged_at") or "")[:10]
        for s in b.get("steps", []):
            target = s.get("target_g")
            actual = s.get("actual_g")
            dev = None if (target is None or actual is None) else actual - target
            verified = s.get("verified")
            vtext = {True: "yes", False: "DISPUTED", None: "not checkable"}.get(
                verified, "")
            # An assumed figure is the target, not a measurement — inventory
            # must be able to tell the two apart.
            measured = "assumed" if s.get("assumed") else "measured"
            cells = [day, b.get("batch_no"), b.get("product"), b.get("base"),
                     b.get("base_weight_g"), s.get("name"), target, actual, dev,
                     s.get("weighed_on"), measured, vtext, b.get("water_ratio")]
            for i, v in enumerate(cells, start=1):
                c = ws.cell(row=row, column=i, value=v)
                if verified is False and i == 12:
                    c.font = RED
                elif s.get("assumed") and i == 11:
                    c.font = MUTED
                else:
                    c.font = BODY
                if i in (5, 7, 8, 9, 13):
                    c.number_format = "#,##0.00" if i in (7, 8, 9) else "#,##0.000"
            row += 1

    if row == 5:
        ws.cell(row=5, column=1, value="No batches logged yet.").font = MUTED
    return row - 5


def build_batches(wb, batches):
    ws = wb.create_sheet("Batches")
    ws["A1"] = "Batches"
    ws["A1"].font = TITLE
    header(ws, 4,
           ["Date", "Batch", "Product", "Base", "Base weight (g)",
            "Ingredients", "Added (g)", "Batch total (g)", "Water ratio",
            "Rebalanced", "Reconciled", "Difference (g)"],
           [12, 10, 16, 10, 14, 12, 12, 14, 11, 12, 12, 13])

    for i, b in enumerate(batches):
        r = 5 + i
        steps = b.get("steps", [])
        added = sum(s.get("actual_g") or 0 for s in steps)
        base = b.get("base_weight_g") or 0
        rec = b.get("reconciliation") or {}
        ok = rec.get("ok")
        cells = [b.get("production_day") or (b.get("logged_at") or "")[:10],
                 b.get("batch_no"), b.get("product"), b.get("base"), base,
                 len(steps), added, base + added, b.get("water_ratio"),
                 "yes" if b.get("rebalanced") else "",
                 {True: "yes", False: "NO"}.get(ok, "n/a"),
                 rec.get("difference_g")]
        for j, v in enumerate(cells, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = RED if (ok is False and j == 11) else BODY
            if j in (5, 7, 8, 12):
                c.number_format = "#,##0.0"
            if j == 9:
                c.number_format = "#,##0.000"
    if not batches:
        ws.cell(row=5, column=1, value="No batches logged yet.").font = MUTED


def build_by_ingredient(wb, batches):
    """The inventory hook: total consumed per ingredient per month."""
    ws = wb.create_sheet("By ingredient")
    ws["A1"] = "Total consumed per ingredient"
    ws["A1"].font = TITLE
    ws["A2"] = "Actual weights, summed by month. This is what inventory draws on."
    ws["A2"].font = MUTED

    totals = defaultdict(float)
    counts = defaultdict(int)
    months = set()
    for b in batches:
        day = b.get("production_day") or (b.get("logged_at") or "")[:10]
        month = day[:7]
        months.add(month)
        for s in b.get("steps", []):
            if s.get("actual_g") is None:
                continue
            totals[(s["name"], month)] += s["actual_g"]
            counts[(s["name"], month)] += 1

    months = sorted(months)
    header(ws, 4, ["Ingredient"] + months + ["Total (g)", "Total (kg)"],
           [26] + [13] * len(months) + [13, 12])

    names = sorted({n for n, _ in totals})
    for i, name in enumerate(names):
        r = 5 + i
        ws.cell(row=r, column=1, value=name).font = BODY
        for j, m in enumerate(months, start=2):
            v = totals.get((name, m))
            if v is not None:
                c = ws.cell(row=r, column=j, value=round(v, 2))
                c.number_format = "#,##0.00"
                c.font = BODY
        first, last = get_column_letter(2), get_column_letter(1 + len(months))
        t = ws.cell(row=r, column=2 + len(months),
                    value=f"=SUM({first}{r}:{last}{r})")
        t.number_format = "#,##0.00"
        t.font = BOLD
        k = ws.cell(row=r, column=3 + len(months),
                    value=f"={get_column_letter(2 + len(months))}{r}/1000")
        k.number_format = "#,##0.000"
        k.font = BODY

    if not names:
        ws.cell(row=5, column=1, value="No batches logged yet.").font = MUTED
    return len(names)


def build(jsonl, out):
    batches = load(jsonl)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    rows = build_consumption(wb, batches)
    build_batches(wb, batches)
    ingredients = build_by_ingredient(wb, batches)
    wb.save(out)
    return len(batches), rows, ingredients


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("jsonl", nargs="?", default=os.path.join(HERE, "batches.jsonl"))
    ap.add_argument("out", nargs="?", default=os.path.join(HERE, "consumption.xlsx"))
    args = ap.parse_args(argv)

    batches, rows, ingredients = build(args.jsonl, args.out)
    print(f"{batches} batch(es), {rows} ingredient rows, "
          f"{ingredients} distinct ingredients -> {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
