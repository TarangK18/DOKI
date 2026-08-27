#!/usr/bin/env python3
"""Build DOKI-Recipes.xlsx — one sheet per recipe, weights in g per 1 kg of meat.

Every derived column uses the same rules the station firmware uses, so what
the sheet says about tolerance and scale routing is what the panel will do.
"""

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "DOKI-Recipes.xlsx"

FONT = "Arial"
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")     # type here
HEAD_FILL = PatternFill("solid", fgColor="1F3552")
TITLE_FILL = PatternFill("solid", fgColor="EEF2F7")
EXAMPLE_FILL = PatternFill("solid", fgColor="F2F2F2")
BLUE = Font(name=FONT, size=10, color="0000FF")          # hardcoded input
BLACK = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
HEAD_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, size=14, bold=True, color="1F3552")
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="666666")
EX_FONT = Font(name=FONT, size=10, italic=True, color="808080")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FIRST_ROW, LAST_ROW = 9, 48          # ingredient rows
TOTAL_ROW = 50

# Teriyaki, from "Recipes for Yield calculation - Teriyaki.csv" (its column 6,
# which is already grams per 1 kg of meat). Spellings normalised; see the
# Instructions sheet.
TERIYAKI = [
    ("Teriyaki sauce", 73.0), ("Soy sauce", 46.4), ("Honey", 42.0),
    ("Tamarind pulp", 25.4), ("White vinegar", 25.4), ("Jaggery powder", 23.2),
    ("Sea salt", 7.5), ("Sesame seed", 4.5), ("Verdad", 2.0),
    ("Onion powder", 1.4), ("Chilli flake", 1.3), ("Black pepper", 1.1),
    ("Yeast extract", 1.0), ("Garlic powder", 0.6), ("Bhut jholokia", 0.25),
    ("Liquid smoke", 0.0),
]

RECIPES = ["Vinegar Bath", "Teriyaki Jerky", "Gochujangh Jerky", "Pepper Jerky",
           "Karnatka Nati Jerky", "Kerala fry Jerky", "Mughlai Jerky",
           "Masala Jerky"]

# Which animal each product is made from. The operator is not asked — the
# product implies it. Blank means not decided yet.
MEAT = {"Karnatka Nati Jerky": "Country chicken"}

# Water is not a recipe percentage — it follows the day's flour, at a ratio the
# supervisor sets each morning. Naming both here is what turns the gate on.
FLOUR_WATER = {
    "Vinegar Bath":  ("", ""),
    "Teriyaki Jerky": ("", ""),    # no flour in the marinade
}

# Master ingredient list: everything Teriyaki uses, plus the ingredients
# already in recipes.json, plus common ones the other six will likely need.
INGREDIENTS = sorted({
    "Teriyaki sauce", "Soy sauce", "Honey", "Tamarind pulp", "White vinegar",
    "Jaggery powder", "Sea salt", "Sesame seed", "Verdad", "Onion powder",
    "Chilli flake", "Black pepper", "Yeast extract", "Garlic powder",
    "Bhut jholokia", "Liquid smoke",
    "Binder (starch)", "Masala spice mix", "Black pepper mix",
    "Classic spice mix", "Schezwan paste", "Salt", "Ice water", "Oil",
    "Rusk / breadcrumb", "Onion paste", "Ginger-garlic paste",
    "Gochujang paste", "Coriander powder", "Cumin powder", "Turmeric",
    "Red chilli powder", "Garam masala", "Curry leaf", "Coconut oil",
    "Mustard seed", "Fennel powder", "Cardamom powder", "Clove powder",
    "Cinnamon powder", "Ginger paste", "Garlic paste", "Lemon juice",
    "Vinegar", "Sugar", "Water",
})

# Cross-sheet references into Settings.
PCT = "Settings!$B$4"
MAIN_DIV = "Settings!$B$7"
SMALL_DIV = "Settings!$B$12"
SMALL_USABLE = "Settings!$B$14"
CROSSOVER = "Settings!$B$19"
MAIN_NAME = "Settings!$B$6"
SMALL_NAME = "Settings!$B$11"


def style_row(ws, row, cols, font=BLACK, fill=None, border=True):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = font
        if fill:
            cell.fill = fill
        if border:
            cell.border = BOX


def sheet_title(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = NOTE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)


# ----------------------------------------------------------------- settings

def build_settings(wb):
    ws = wb.create_sheet("Settings")
    sheet_title(ws, "Settings",
                "These drive every recipe sheet. Change a number here and all "
                "seven recipes re-evaluate.")

    rows = [
        (4, "Tolerance (percent of target)", 0.02, "0.0%",
         "What the recipe asks for. The station uses this too."),
        (5, "", None, None, None),
        (6, "MAIN SCALE — name", "Floor scale", None, ""),
        (7, "  division (g)", 5, "0.0",
         "Smallest step it reads. From the RS232 frame resolution."),
        (8, "  capacity (g)", 50000, "#,##0", ""),
        (9, "  usable (g)", 50000, "#,##0", "Capacity less the tub's tare."),
        (10, "", None, None, None),
        (11, "SMALL SCALE — name", "Bench scale", None, ""),
        (12, "  division (g)", 0.1, "0.00",
         "SPEC NOT CONFIRMED — check the label on the scale."),
        (13, "  capacity (g)", 3000, "#,##0", "SPEC NOT CONFIRMED."),
        (14, "  usable (g)", 2000, "#,##0",
         "Capacity less the tare of the container you weigh into."),
    ]
    for r, labelled, value, fmt, note in rows:
        if labelled == "":
            continue
        ws.cell(row=r, column=1, value=labelled).font = (
            BOLD if labelled.endswith("name") or "Tolerance" in labelled else BLACK)
        c = ws.cell(row=r, column=2, value=value)
        c.font = BLUE
        c.fill = INPUT_FILL
        c.border = BOX
        if fmt:
            c.number_format = fmt
        if note:
            ws.cell(row=r, column=3, value=note).font = NOTE_FONT

    ws["A17"] = "Crossover — derived"
    ws["A17"].font = BOLD
    ws["B17"] = f"=2*{MAIN_DIV}/{PCT}"
    ws["B17"].number_format = "#,##0"
    ws["B17"].border = BOX
    ws["C17"] = ("Below this the main scale cannot hold the tolerance: two of "
                 "its divisions no longer fit inside the percentage.")
    ws["C17"].font = NOTE_FONT

    ws["A18"] = "Crossover — override"
    ws["B18"] = None
    ws["B18"].fill = INPUT_FILL
    ws["B18"].font = BLUE
    ws["B18"].border = BOX
    ws["B18"].number_format = "#,##0"
    ws["C18"] = ("Optional. Leave blank to use the derived value. A lower "
                 "number keeps more ingredients on the floor scale — easier to "
                 "pour into, looser tolerance, and the recipe sheets say which.")
    ws["C18"].font = NOTE_FONT

    ws["A19"] = "Crossover in force (g)"
    ws["A19"].font = BOLD
    ws["B19"] = '=IF(B18="",B17,B18)'
    ws["B19"].font = BOLD
    ws["B19"].number_format = "#,##0"
    ws["B19"].border = BOX

    ws["A21"] = "Dead zone check"
    ws["A21"].font = BOLD
    ws["B21"] = (f'=IF({SMALL_USABLE}>={CROSSOVER},"None — the two scales meet",'
                 f'"WARNING: nothing can weigh "&TEXT({SMALL_USABLE},"#,##0")&'
                 f'" to "&TEXT({CROSSOVER},"#,##0")&" g")')
    ws["B21"].font = BOLD
    ws.merge_cells("B21:E21")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 70
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


# -------------------------------------------------------------- ingredients

def build_ingredients(wb):
    ws = wb.create_sheet("Ingredients")
    sheet_title(ws, "Ingredient list",
                "Feeds the dropdown on every recipe sheet. Add new names at "
                "the bottom — keep one spelling per ingredient.")
    ws["A4"] = "Ingredient"
    ws["A4"].font = HEAD_FONT
    ws["A4"].fill = HEAD_FILL
    ws["A4"].border = BOX
    for i, name in enumerate(INGREDIENTS):
        c = ws.cell(row=5 + i, column=1, value=name)
        c.font = BLACK
        c.border = BOX
    ws["C5"] = ("Used in:")
    ws["C5"].font = BOLD
    for i, name in enumerate(INGREDIENTS):
        refs = " + ".join(
            f'COUNTIF(\'{r}\'!$B${FIRST_ROW}:$B${LAST_ROW},$A{5 + i})'
            for r in RECIPES)
        c = ws.cell(row=5 + i, column=3, value=f"={refs}")
        c.font = BLACK
        c.number_format = "0"
        c.border = BOX
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["C"].width = 10
    return ws


# ------------------------------------------------------------------ recipes

def build_recipe(wb, name, data=None, example=False):
    ws = wb.create_sheet(name)
    sheet_title(ws, name,
                "Type in the shaded cells only. Weights are GRAMS PER 1 KG OF "
                "MEAT. Everything to the right is calculated.")

    # Labels merged across A:B so they are not clipped by the input beside them.
    ws.merge_cells("A4:B4")
    ws["A4"] = "Batch base weight (g)"
    ws["A4"].font = BOLD
    ws["A4"].alignment = Alignment(horizontal="right")
    ws["C4"] = 3200
    ws["C4"].number_format = "#,##0"
    ws.merge_cells("D4:H4")
    ws["D4"] = ("A typical batch. Only the Target / Weigh-on / Tolerance columns "
                "use it — the station recomputes them from the meat it actually "
                "weighs.")
    ws["D4"].font = NOTE_FONT

    ws.merge_cells("A5:B5")
    ws["A5"] = "Product ID"
    ws["A5"].font = BOLD
    ws["A5"].alignment = Alignment(horizontal="right")
    ws["C5"] = name.lower().replace(" ", "_")

    ws.merge_cells("D5:E5")
    ws["D5"] = "Bases this applies to"
    ws["D5"].font = BOLD
    ws["D5"].alignment = Alignment(horizontal="right")
    ws.merge_cells("F5:H5")
    ws["F5"] = "chicken, pork, beef, mutton, fish"

    # Blank by default: naming a flour that is not in the sheet would fail the
    # check below, and none of these recipes is known to use one yet.
    flour, water = FLOUR_WATER.get(name, ("", ""))
    ws.merge_cells("A7:B7")
    ws["A7"] = "Meat"
    ws["A7"].font = BOLD
    ws["A7"].alignment = Alignment(horizontal="right")
    ws["C7"] = MEAT.get(name, "")
    ws["C7"].font = BLUE
    ws["C7"].fill = INPUT_FILL
    ws["C7"].border = BOX
    ws.merge_cells("D7:H7")
    ws["D7"] = ("Which animal this product is made from. The operator is not "
                "asked — the product implies it. Leave blank if undecided.")
    ws["D7"].font = NOTE_FONT

    ws.merge_cells("A6:B6")
    ws["A6"] = "Flour ingredient"
    ws["A6"].font = BOLD
    ws["A6"].alignment = Alignment(horizontal="right")
    ws["C6"] = flour
    ws.merge_cells("D6:E6")
    ws["D6"] = "Water ingredient"
    ws["D6"].font = BOLD
    ws["D6"].alignment = Alignment(horizontal="right")
    ws.merge_cells("F6:H6")
    ws["F6"] = water

    for ref in ("C4", "C5", "F5", "C6", "F6"):
        ws[ref].font = BLUE
        ws[ref].fill = INPUT_FILL
        ws[ref].border = BOX
    ws["C4"].alignment = Alignment(horizontal="right")

    # Both must name a row in this sheet, or neither. A half-specified pair
    # means water cannot be derived and the converter will refuse the recipe.
    check = ws.cell(row=53, column=2, value=(
        f'=IF(AND($C$6="",$F$6=""),"Not water-gated — water comes from the '
        f'recipe row as usual.",'
        f'IF(OR($C$6="",$F$6=""),"ERROR: name BOTH a flour and a water '
        f'ingredient, or neither.",'
        f'IF(COUNTIF($B${FIRST_ROW}:$B${LAST_ROW},$C$6)=0,'
        f'"ERROR: flour ingredient """&$C$6&""" is not in the list below.",'
        f'IF(COUNTIF($B${FIRST_ROW}:$B${LAST_ROW},$F$6)=0,'
        f'"ERROR: water ingredient """&$F$6&""" is not in the list below.",'
        f'"Water is set daily by the supervisor: "&$F$6&" = ratio x "&$C$6&". '
        f'The weight typed against "&$F$6&" below is only the usual figure.")))'
        f')'))
    check.font = BOLD
    ws.merge_cells(start_row=53, start_column=2, end_row=53, end_column=8)
    ws.conditional_formatting.add("B53:H53", FormulaRule(
        formula=['LEFT($B$53,5)="ERROR"'],
        fill=PatternFill("solid", fgColor="FBD5D5"),
        font=Font(name=FONT, size=10, bold=True, color="9C0006")))

    headers = ["#", "Ingredient", "g per 1 kg meat", "% of meat",
               f"Target in this batch", "Weigh on", "Tolerance ±", "Note"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=8, column=i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for r in range(FIRST_ROW, LAST_ROW + 1):
        i = r - FIRST_ROW
        ws.cell(row=r, column=1, value=f'=IF($B{r}="","",{i + 1})')

        # --- inputs
        b = ws.cell(row=r, column=2)
        c = ws.cell(row=r, column=3)
        for cell in (b, c):
            cell.fill = INPUT_FILL
            cell.font = BLUE
            cell.border = BOX
        c.number_format = "0.00"

        if data and i < len(data):
            b.value, c.value = data[i]
        elif example and i == 0:
            b.value, c.value = "Salt", 7.5
            b.font = c.font = EX_FONT
            b.fill = c.fill = EXAMPLE_FILL

        # --- derived
        pct = ws.cell(row=r, column=4, value=f'=IF($B{r}="","",$C{r}/1000)')
        pct.number_format = "0.000%"

        tgt = ws.cell(row=r, column=5, value=f'=IF($B{r}="","",$D{r}*$C$4)')
        tgt.number_format = "#,##0.00"

        # Same rule as the firmware: the main scale only holds the percentage
        # once two of its divisions fit inside it.
        ws.cell(row=r, column=6, value=(
            f'=IF($B{r}="","",'
            f'IF($E{r}<=0,"NEITHER — zero",'
            f'IF($E{r}>={CROSSOVER},{MAIN_NAME},'
            f'IF($E{r}>{SMALL_USABLE},"NEITHER — over bench capacity",'
            f'IF($E{r}>=2*{SMALL_DIV},{SMALL_NAME},'
            f'"NEITHER — under bench resolution")))))'))

        tol = ws.cell(row=r, column=7, value=(
            f'=IF($B{r}="","",'
            f'IF($F{r}={MAIN_NAME},MAX(2*{MAIN_DIV},{PCT}*$E{r}),'
            f'IF($F{r}={SMALL_NAME},MAX(2*{SMALL_DIV},{PCT}*$E{r}),"")))'))
        tol.number_format = "#,##0.00"

        ws.cell(row=r, column=8, value=(
            f'=IF($B{r}="","",'
            f'IF(LEFT($F{r},7)="NEITHER","Cannot be weighed — split it, premix '
            f'it, or raise the batch size",'
            f'IF($G{r}>{PCT}*$E{r}+0.0001,'
            f'"Held to the scale\'s resolution, not "&TEXT({PCT},"0%"),"")))'))

        style_row(ws, r, range(1, 9))
        ws.cell(row=r, column=1).font = BLACK
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        for col in (4, 5, 7):
            ws.cell(row=r, column=col).font = BLACK
        ws.cell(row=r, column=6).font = BLACK
        ws.cell(row=r, column=8).font = NOTE_FONT
        if data and i < len(data):
            pass
        elif example and i == 0:
            ws.cell(row=r, column=8).font = EX_FONT

    # --- totals
    ws.cell(row=TOTAL_ROW, column=2, value="TOTAL").font = BOLD
    t = ws.cell(row=TOTAL_ROW, column=3,
                value=f"=SUM(C{FIRST_ROW}:C{LAST_ROW})")
    t.number_format = "#,##0.00"
    p = ws.cell(row=TOTAL_ROW, column=4,
                value=f"=SUM(D{FIRST_ROW}:D{LAST_ROW})")
    p.number_format = "0.000%"
    g = ws.cell(row=TOTAL_ROW, column=5,
                value=f"=SUM(E{FIRST_ROW}:E{LAST_ROW})")
    g.number_format = "#,##0.00"
    for col in (2, 3, 4, 5):
        ws.cell(row=TOTAL_ROW, column=col).font = BOLD
        ws.cell(row=TOTAL_ROW, column=col).border = BOX

    ws.cell(row=TOTAL_ROW + 1, column=2, value="Batch total with meat (g)").font = BOLD
    bt = ws.cell(row=TOTAL_ROW + 1, column=5, value=f"=$C$4+E{TOTAL_ROW}")
    bt.number_format = "#,##0.00"
    bt.font = BOLD

    ws.cell(row=TOTAL_ROW + 3, column=2, value=(
        f'=IF(COUNTIF($F${FIRST_ROW}:$F${LAST_ROW},"NEITHER*")=0,'
        f'"All ingredients can be weighed.",'
        f'"WARNING: "&COUNTIF($F${FIRST_ROW}:$F${LAST_ROW},"NEITHER*")&'
        f'" ingredient(s) cannot be weighed on either scale — see the Note column.")')
        ).font = BOLD
    ws.merge_cells(start_row=TOTAL_ROW + 3, start_column=2,
                   end_row=TOTAL_ROW + 3, end_column=8)

    # --- dropdown, free text still allowed
    dv = DataValidation(type="list", formula1="=IngredientList",
                        allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv)
    dv.add(f"B{FIRST_ROW}:B{LAST_ROW}")

    # --- conditional formatting
    rng = f"A{FIRST_ROW}:H{LAST_ROW}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'LEFT($F{FIRST_ROW},7)="NEITHER"'],
        fill=PatternFill("solid", fgColor="FBD5D5"),
        font=Font(name=FONT, size=10, color="9C0006")))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND($B{FIRST_ROW}<>"",$H{FIRST_ROW}<>"")'],
        fill=PatternFill("solid", fgColor="FFF3CD")))

    widths = {"A": 5, "B": 26, "C": 16, "D": 11, "E": 16, "F": 18, "G": 13, "H": 46}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{FIRST_ROW}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:H{TOTAL_ROW + 4}"
    return ws


# ------------------------------------------------------------------ summary

def build_summary(wb):
    ws = wb.create_sheet("Summary")
    sheet_title(ws, "All recipes",
                "Rolls up every recipe sheet. Nothing to type here.")
    headers = ["Recipe", "Ingredients", "Total g per 1 kg meat", "Total % of meat",
               "On floor scale", "On bench scale", "Cannot be weighed"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, name in enumerate(RECIPES):
        r = 5 + i
        q = f"'{name}'"
        ws.cell(row=r, column=1, value=name).font = BOLD
        ws.cell(row=r, column=2,
                value=f"=COUNTA({q}!$B${FIRST_ROW}:$B${LAST_ROW})")
        ws.cell(row=r, column=3, value=f"={q}!$C${TOTAL_ROW}").number_format = "#,##0.00"
        ws.cell(row=r, column=4, value=f"={q}!$D${TOTAL_ROW}").number_format = "0.000%"
        ws.cell(row=r, column=5, value=(
            f"=COUNTIF({q}!$F${FIRST_ROW}:$F${LAST_ROW},{MAIN_NAME})"))
        ws.cell(row=r, column=6, value=(
            f"=COUNTIF({q}!$F${FIRST_ROW}:$F${LAST_ROW},{SMALL_NAME})"))
        ws.cell(row=r, column=7, value=(
            f'=COUNTIF({q}!$F${FIRST_ROW}:$F${LAST_ROW},"NEITHER*")'))
        style_row(ws, r, range(1, 8))
        ws.cell(row=r, column=1).font = BOLD

    widths = {"A": 18, "B": 13, "C": 20, "D": 16, "E": 15, "F": 15, "G": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


# ------------------------------------------------------------- instructions

def build_instructions(wb):
    ws = wb.create_sheet("Instructions", 0)
    sheet_title(ws, "DOKI recipe workbook",
                "One sheet per recipe. Weights are grams per 1 kg of meat.")

    lines = [
        ("How to use", True),
        ("Open the sheet for a recipe. In the shaded columns type the ingredient "
         "name and how many grams of it go into 1 kg of meat. That is all.", False),
        ("Everything to the right calculates itself: the percentage, the target "
         "for a typical batch, which scale should weigh it, and the tolerance "
         "that scale can actually hold.", False),
        ("", False),
        ("What the colours mean", True),
        ("Shaded cells are yours to type in. Everything else is a formula — "
         "overwriting one breaks that row.", False),
        ("A red row means neither scale can weigh that ingredient at this batch "
         "size. Split it, premix it, or make the batch bigger.", False),
        ("An amber row is fine to run, but the scale's resolution is coarser than "
         "the recipe's percentage, so it is held to a looser band. The Note "
         "column says so.", False),
        ("", False),
        ("Which scale gets what", True),
        ("The floor scale reads in 5 g steps. Two of those have to fit inside the "
         "tolerance before it is enforcing anything, so it only handles targets "
         "above 500 g — that is 2 × 5 g ÷ 2 %, on the Settings sheet. Everything "
         "smaller goes on the bench scale.", False),
        ("Change the tolerance percentage or either scale's division on the "
         "Settings sheet and all seven recipes re-evaluate.", False),
        ("", False),
        ("About the Teriyaki sheet", True),
        ("Filled in from 'Recipes for Yield calculation - Teriyaki.csv', using its "
         "column 6, which is already grams per 1 kg of meat.", False),
        ("Two spellings were normalised so the ingredient list stays consistent: "
         "'seasame seed' → 'Sesame seed', and capitalisation made uniform.", False),
        ("Liquid smoke is in the source sheet with a quantity of 0, so it shows "
         "as unweighable here. Either give it a quantity or delete the row.", False),
        ("The very small spices are below what a floor scale can see. The costing "
         "sheet treats the marinade as one premixed line at 255 g per kg of meat, "
         "and that is how the station should weigh it too — premix the marinade "
         "on the bench scale, then add it as a single ingredient.", False),
        ("", False),
        ("The other six recipes", True),
        ("Gochujang, Pepper, Nati, Kerela Fry, Mughlai and Masala are set up but "
         "empty — the formulas are already in every row. Each has one grey example "
         "row showing the format; type over it.", False),
        ("", False),
        ("Feeding the station", True),
        ("Run  python3 xlsx_to_recipes.py DOKI-Recipes.xlsx  to turn this workbook "
         "into the recipes.json the panel reads. It refuses to write a recipe that "
         "still has an unweighable ingredient or an untouched example row.", False),
    ]
    r = 4
    for text, is_head in lines:
        if not text:
            r += 1
            continue
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name=FONT, size=11, bold=True, color="1F3552") if is_head \
            else Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        if not is_head:
            ws.row_dimensions[r].height = 28
        r += 1

    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 18
    return ws


# --------------------------------------------------------------------- main

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_instructions(wb)
    build_settings(wb)
    for name in RECIPES:
        build_recipe(wb, name,
                     data=TERIYAKI if name == "Teriyaki" else None,
                     example=(name != "Teriyaki"))
    build_summary(wb)
    ing = build_ingredients(wb)

    wb.defined_names.add(DefinedName(
        "IngredientList",
        attr_text=f"Ingredients!$A$5:$A${4 + len(INGREDIENTS)}"))

    wb.active = 0
    wb.save(OUT)
    print(f"wrote {OUT}: {len(wb.sheetnames)} sheets — {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
