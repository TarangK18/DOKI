#!/usr/bin/env python3
"""Turn DOKI-Recipes.xlsx into the recipes.json the station reads.

  python3 xlsx_to_recipes.py DOKI-Recipes.xlsx [recipes.json]

The workbook holds grams per 1 kg of meat; recipes.json holds a percentage of
the base weight. 73 g/kg is 7.3 %.

Refuses to write a recipe that still has an unweighable ingredient or an
untouched example row, so a sheet nobody has finished cannot quietly reach the
floor. Existing bases, scales, tolerances and the PIN are left alone.
"""

import argparse
import json
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
FIRST_ROW, LAST_ROW = 8, 47
SKIP = {"Instructions", "Settings", "Summary", "Ingredients"}
EXAMPLE = ("Salt", 7.5)      # the grey row seeded into each blank sheet


def read_sheet(ws):
    """(product_dict, [problems]) for one recipe sheet."""
    problems = []
    product_id = (ws["C5"].value or "").strip()
    bases = [b.strip() for b in str(ws["F5"].value or "").split(",") if b.strip()]
    if not product_id:
        problems.append("no Product ID in C5")
    if not bases:
        problems.append("no bases listed in F5")

    flour = (ws["C6"].value or "").strip()
    water = (ws["F6"].value or "").strip()
    if bool(flour) != bool(water):
        problems.append("names a flour or a water ingredient but not both — "
                        "water cannot be derived from the daily ratio")

    ingredients, seen = [], set()
    for r in range(FIRST_ROW, LAST_ROW + 1):
        name = ws.cell(row=r, column=2).value
        if name is None or not str(name).strip():
            continue
        name = str(name).strip()
        grams = ws.cell(row=r, column=3).value
        where = ws.cell(row=r, column=6).value or ""

        if grams is None:
            problems.append(f"row {r}: '{name}' has no weight")
            continue
        if str(where).startswith("NEITHER"):
            problems.append(f"row {r}: '{name}' — {where}")
            continue
        if name.lower() in seen:
            problems.append(f"row {r}: '{name}' appears twice")
            continue
        seen.add(name.lower())
        ingredients.append([name, round(float(grams) / 10.0, 4)])

    if not ingredients:
        problems.append("no ingredients filled in")
    elif len(ingredients) == 1 and ingredients[0][0] == EXAMPLE[0] \
            and abs(ingredients[0][1] - EXAMPLE[1] / 10.0) < 1e-9:
        problems.append("still holds only the grey example row")

    names = {n for n, _ in ingredients}
    for role, ing in (("flour", flour), ("water", water)):
        if ing and ing not in names:
            problems.append(f"{role} ingredient '{ing}' is not in the "
                            f"ingredient list")

    product = {"id": product_id, "name": ws.title, "bases": bases,
               "ingredients": ingredients}
    if flour and water:
        product["flour_ingredient"] = flour
        product["water_ingredient"] = water
    return product, problems


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook", nargs="?",
                    default=os.path.join(HERE, "DOKI-Recipes.xlsx"))
    ap.add_argument("recipes", nargs="?",
                    default=os.path.join(HERE, "recipes.json"))
    ap.add_argument("--dry-run", action="store_true",
                    help="report what would change without writing")
    args = ap.parse_args(argv)

    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    with open(args.recipes, "r", encoding="utf-8") as fh:
        cfg = json.load(fh)

    ready, skipped = [], []
    for title in wb.sheetnames:
        if title in SKIP:
            continue
        product, problems = read_sheet(wb[title])
        if problems:
            skipped.append((title, problems))
        else:
            ready.append(product)

    for title, problems in skipped:
        print(f"skipped {title}:")
        for p in problems:
            print(f"    {p}")

    if not ready:
        print("\nNothing to write — no sheet is complete.")
        return 1

    by_id = {p["id"]: p for p in cfg.get("products", [])}
    added, updated = [], []
    for p in ready:
        (updated if p["id"] in by_id else added).append(p["name"])
        by_id[p["id"]] = p
    cfg["products"] = list(by_id.values())

    print(f"\nready: {len(ready)} recipe(s)")
    for p in ready:
        total = sum(pct for _, pct in p["ingredients"])
        gate = (f"  water = ratio x {p['flour_ingredient']}"
                if p.get("flour_ingredient") else "  not water-gated")
        print(f"    {p['name']:<14} {len(p['ingredients']):>2} ingredients, "
              f"{total:.3f} % of base ({total * 10:.2f} g per kg of meat)"
              f"{gate}")
    if added:
        print(f"  added:   {', '.join(added)}")
    if updated:
        print(f"  updated: {', '.join(updated)}")

    if args.dry_run:
        print("\n--dry-run: recipes.json not written")
        return 0

    with open(args.recipes, "w", encoding="utf-8") as fh:
        json.dump(cfg, fh, indent=2, ensure_ascii=False)
        fh.write("\n")
    print(f"\nwrote {args.recipes}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
